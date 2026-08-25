"""
extract_kpis.py

Reconstrucao do script que extrai os KPIs regulatorios da planilha
"LISTA MESTRE DE PRODUTOS - PRONUTRITION.xlsx" e gera o dict `D` usado pelo
dashboard HTML (Dashboard Regulatorios - RASCUNHO Novos Indicadores.html).

So usa openpyxl (read_only, data_only) + stdlib. Cabecalhos nao estao
necessariamente na primeira linha de cada aba -- localizamos a linha de
cabecalho procurando pelo conteudo esperado (nome de pelo menos uma coluna
conhecida).

Uso:
    from extract_kpis import build
    d = build("LISTA MESTRE DE PRODUTOS - PRONUTRITION.xlsx")
"""

import re
import statistics
from collections import defaultdict, Counter
from datetime import datetime, date


# --------------------------------------------------------------------------
# Normalizacao de nomes (correcoes de digitacao conhecidas na planilha)
# --------------------------------------------------------------------------
_NAME_FIXES = {
    "pronutrion": "Pronutrition",
    "pronutriton": "Pronutrition",
    "pronutrition": "Pronutrition",
    "decatlhon": "Decathlon",
    "decathlon": "Decathlon",
    "talita tozo": "Talita Tozzo",
    "talita tozzo": "Talita Tozzo",
}

# Nomes conhecidos, usados para "title-case" bonito quando a planilha traz
# tudo em CAIXA ALTA (aba 'Gastos Reais').
_KNOWN_PROPER = {
    "AJINOMOTO": "Ajinomoto",
    "CHACOMER": "Chacomer",
    "CIMED": "Cimed",
    "DECATHLON": "Decathlon",
    "FEPASE": "Fepase",
    "HUMANA": "Humana",
    "PRONUTRITION": "Pronutrition",
    "SUNRIZE": "Sunrize",
    "VITAMIDAS": "Vitamidas",
    "TALITA TOZO": "Talita Tozzo",
    "TALITA TOZZO": "Talita Tozzo",
}


def normalize_name(value):
    """Normaliza nomes de cliente/marca/laboratorio corrigindo erros de
    digitacao conhecidos ('Pronutrion'/'Pronutriton' -> 'Pronutrition',
    'DECATLHON' -> 'Decathlon', variacoes de 'Talita Tozzo' etc.) e
    padronizando capitalizacao de nomes que vierem em CAIXA ALTA.
    """
    if value is None:
        return value
    s = str(value).replace("\xa0", " ").strip()
    if not s:
        return s
    key = re.sub(r"\s+", " ", s).strip().lower()
    if key in _NAME_FIXES:
        return _NAME_FIXES[key]
    upper = re.sub(r"\s+", " ", s).strip().upper()
    if upper in _KNOWN_PROPER:
        return _KNOWN_PROPER[upper]
    if s.isupper() and len(s) > 1:
        return s.title()
    return re.sub(r"\s+", " ", s).strip()


# --------------------------------------------------------------------------
# Helpers genericos
# --------------------------------------------------------------------------

def _clean(v):
    if isinstance(v, str):
        v = v.replace("\xa0", " ").strip()
        return v
    return v


def _find_header_row(ws, expected_cols, max_scan=15):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= max_scan:
            break
    best_idx, best_score = 0, -1
    for i, row in enumerate(rows):
        cells = [str(_clean(c)) for c in row if c is not None]
        score = sum(1 for exp in expected_cols if any(exp.lower() in c.lower() for c in cells))
        if score > best_score:
            best_score = score
            best_idx = i
    header = [_clean(c) for c in rows[best_idx]]
    return best_idx, header


def _col_index(header, *names):
    for name in names:
        for i, h in enumerate(header):
            if h and name.lower() in str(h).lower():
                return i
    raise ValueError(f"Coluna nao encontrada para {names} em {header}")


def _iter_data_rows(ws, header_idx):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        if row is None or all(c is None for c in row):
            continue
        yield [_clean(c) for c in row]


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                continue
    return None


def _to_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("R$", "").replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


MESES_PT = [
    "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]
# nomes com acento para saida (o mes de indice 2 precisa ser "marco" com cedilha)
MESES_PT_ACENTO = [
    "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]


def _month_name(d):
    return MESES_PT_ACENTO[d.month - 1]


# --------------------------------------------------------------------------
# 1) Cotacoes
# --------------------------------------------------------------------------

def _extract_cotacoes(wb):
    ws = wb["Gestão de Cotações"]
    header_idx, header = _find_header_row(ws, ["Aprovado", "Solicitação", "Recebimento", "Valor"])
    c_status = _col_index(header, "Aprovado")
    c_sol = _col_index(header, "Solicitação")
    c_rec = _col_index(header, "Recebimento")
    c_val = _col_index(header, "Valor")

    status_count = Counter()
    valores_sum = defaultdict(float)
    leads = []

    for row in _iter_data_rows(ws, header_idx):
        status = row[c_status] if c_status < len(row) else None
        if status is None or str(status).strip() == "":
            continue
        status = str(status).strip()
        status_count[status] += 1

        val = _to_number(row[c_val]) if c_val < len(row) else None
        if val is not None:
            valores_sum[status] += val

        d_sol = _to_date(row[c_sol]) if c_sol < len(row) else None
        d_rec = _to_date(row[c_rec]) if c_rec < len(row) else None
        if d_sol and d_rec:
            delta = (d_rec - d_sol).days
            if 0 <= delta <= 365:
                leads.append(delta)

    lead_medio = round(statistics.mean(leads), 2) if leads else 0.0
    return {
        "status": dict(status_count),
        "lead_medio_sol_rec": lead_medio,
        "lead_n": len(leads),
        "valores": {k: round(v, 2) for k, v in valores_sum.items()},
    }


# --------------------------------------------------------------------------
# 2) Fluxo de caixa (Gastos Catalogo)
# --------------------------------------------------------------------------

def _extract_fluxo(wb):
    ws = wb["Gastos Catálogo"]
    header_idx, header = _find_header_row(ws, ["Produto", "Lab", "Data", "Valor Pagamento"])
    c_lab = _col_index(header, "Lab")
    c_valor = _col_index(header, "Valor Pagamento")
    c_data = None
    for i in range(c_valor - 1, -1, -1):
        if header[i] and "data" in str(header[i]).lower():
            c_data = i
            break
    if c_data is None:
        c_data = _col_index(header, "Data")

    meses = defaultdict(lambda: defaultdict(float))
    labs = set()

    for row in _iter_data_rows(ws, header_idx):
        lab = normalize_name(row[c_lab]) if c_lab < len(row) else None
        d = _to_date(row[c_data]) if c_data < len(row) else None
        val = _to_number(row[c_valor]) if c_valor < len(row) else None
        if not lab or not d or val is None:
            continue
        mkey = f"{d.year:04d}-{d.month:02d}"
        meses[mkey][lab] += val
        labs.add(lab)

    meses_out = {k: {lab: round(v, 2) for lab, v in labdict.items()} for k, labdict in sorted(meses.items())}
    return {"meses": meses_out, "labs": sorted(labs)}


# --------------------------------------------------------------------------
# 3) Estabilidade (funil + proximos laudos) + qualidade_lab + prob_prod
# --------------------------------------------------------------------------

def _extract_estab_qualidade_prob(wb, today):
    ws = wb["Controle Estabilidade "]
    header_idx, header = _find_header_row(
        ws, ["Descrição - DATASUL", "Laboratório", "Tempo do estudo", "Previsão de laudo",
             "Recebimento de laudo", "Laudo recebido", "STATUS", "Envio da amostra"]
    )
    c_prod = _col_index(header, "Descrição - DATASUL")
    c_lab = _col_index(header, "Laboratório")
    c_tempo = _col_index(header, "Tempo do estudo")
    c_prev = _col_index(header, "Previsão de laudo")
    c_recebimento = _col_index(header, "Recebimento de laudo")
    c_recebido = _col_index(header, "Laudo recebido")
    c_status = _col_index(header, "STATUS")
    try:
        c_envio = _col_index(header, "Envio da amostra")
    except ValueError:
        c_envio = None

    tempos = Counter()
    proximos = []
    qualidade_lab = defaultdict(Counter)
    lab_sla_acc = defaultdict(lambda: {"n_avaliados": 0, "n_atrasados": 0, "_entregas": []})

    for row in _iter_data_rows(ws, header_idx):
        prod = row[c_prod] if c_prod < len(row) else None
        if not prod:
            continue
        tempo = row[c_tempo] if c_tempo < len(row) else None
        if tempo:
            tempos[str(tempo).strip()] += 1

        lab = row[c_lab] if c_lab < len(row) else None
        status = row[c_status] if c_status < len(row) else None
        lab_norm = str(lab).strip().upper() if lab else None
        if lab_norm and status:
            qualidade_lab[lab_norm][str(status).strip()] += 1

        recebido = row[c_recebido] if c_recebido < len(row) else None
        prev = row[c_prev] if c_prev < len(row) else None
        prev_date = _to_date(prev)
        recebimento_date = _to_date(row[c_recebimento]) if c_recebimento < len(row) else None

        # taxa de atraso e tempo medio de entrega de laudo, por laboratorio:
        # ambos calculados pela diferenca entre "Previsao de laudo" e
        # "Recebimento de laudo" (somente laudos ja recebidos, com as duas
        # datas preenchidas).
        if lab_norm and recebido == "Sim" and prev_date and recebimento_date:
            acc = lab_sla_acc[lab_norm]
            acc["n_avaliados"] += 1
            delta = (recebimento_date - prev_date).days
            if delta > 0:
                acc["n_atrasados"] += 1
            acc["_entregas"].append(delta)

        if recebido != "Sim" and prev_date:
            proximos.append({
                "produto": str(prod).strip(),
                "lab": str(lab).strip().upper() if lab else "",
                "tempo": str(tempo).strip() if tempo else "",
                "previsao": prev_date.isoformat(),
                "_delta": abs((prev_date - today).days),
                "_futuro": prev_date >= today,
            })

    proximos.sort(key=lambda x: (not x["_futuro"], x["_delta"]))
    # "estudos de estabilidade em andamento" conta cada produto uma unica vez
    # (nao o total de laudos pendentes, que pode ter varios por produto).
    n_proximos_total = len({p["produto"] for p in proximos})
    proximos_out = [
        {k: v for k, v in p.items() if not k.startswith("_")}
        for p in proximos[:15]
    ]

    estab = {
        "tempos": dict(tempos),
        "proximos_laudos": proximos_out,
        # n_proximos = total de laudos aguardados (nao apenas os ~15 exibidos)
        "n_proximos": n_proximos_total,
    }

    qualidade_lab_out = {lab: dict(counter) for lab, counter in qualidade_lab.items()}

    lab_sla = {}
    for lab, acc in lab_sla_acc.items():
        n = acc["n_avaliados"]
        entregas = acc["_entregas"]
        lab_sla[lab] = {
            "n_avaliados": n,
            "taxa_atraso": round(100 * acc["n_atrasados"] / n, 1) if n else 0.0,
            "tempo_medio_entrega": round(statistics.mean(entregas), 1) if entregas else None,
        }

    return estab, qualidade_lab_out, lab_sla


# --------------------------------------------------------------------------
# 4) Documentacoes (Controle Documentos) -> docs + rev_arte
# --------------------------------------------------------------------------

def _extract_docs_and_revarte(wb):
    ws = wb["Controle Documentos"]
    header_idx, header = _find_header_row(
        ws, ["Descrição - DATASUL", "Tipo de documento", "DATA ENTRADA", "DATA SAÍDA", "SLA", "VERSÃO", "Cliente"]
    )
    c_prod_datasul = _col_index(header, "Descrição - DATASUL")
    c_cliente = _col_index(header, "Cliente")
    c_marca = _col_index(header, "Marca")
    c_tipo = _col_index(header, "Tipo de documento")
    c_versao = _col_index(header, "VERSÃO")
    c_entrada = _col_index(header, "DATA ENTRADA")
    c_saida = _col_index(header, "DATA SAÍDA")
    c_sla = _col_index(header, "SLA")

    EXCLUDED_TIPOS = {"notificação", "notificacao", "fechamento de arte", "especificação técnica", "especificacao tecnica"}

    meses = defaultdict(lambda: defaultdict(lambda: {"qtd": 0, "_slas": []}))
    tipos_seen = set()

    arte_group = defaultdict(list)

    for row in _iter_data_rows(ws, header_idx):
        prod = row[c_prod_datasul] if c_prod_datasul < len(row) else None
        if not prod:
            continue
        tipo = row[c_tipo] if c_tipo < len(row) else None
        if not tipo:
            continue
        tipo = str(tipo).strip()
        if tipo.lower() in EXCLUDED_TIPOS:
            continue

        entrada = _to_date(row[c_entrada]) if c_entrada < len(row) else None
        sla_raw = row[c_sla] if c_sla < len(row) else None
        sla = _to_number(sla_raw)

        if entrada:
            mkey = _month_name(entrada)
        else:
            mkey = "sem data"

        bucket = meses[mkey][tipo]
        bucket["qtd"] += 1
        tipos_seen.add(tipo)
        if sla is not None and 0 <= sla <= 200:
            bucket["_slas"].append(sla)

        if tipo == "Correção de Arte":
            cliente = normalize_name(row[c_cliente]) if c_cliente < len(row) else None
            marca = normalize_name(row[c_marca]) if c_marca < len(row) else None
            versao = row[c_versao] if c_versao < len(row) else None
            versao_num = None
            if isinstance(versao, (int, float)):
                versao_num = int(versao)
            elif isinstance(versao, str) and versao.strip().isdigit():
                versao_num = int(versao.strip())
            key = (str(prod).strip(), cliente or "")
            arte_group[key].append({
                "marca": marca or "",
                "versao": versao_num,
                "entrada": entrada,
            })

    ordem = [m for m in MESES_PT_ACENTO if m in meses] + (["sem data"] if "sem data" in meses else [])
    tipos = sorted(tipos_seen)

    meses_out = {}
    for mkey, tipos_dict in meses.items():
        meses_out[mkey] = {}
        for tipo, bucket in tipos_dict.items():
            slas = bucket["_slas"]
            sla_avg = round(statistics.mean(slas), 2) if slas else 0.0
            meses_out[mkey][tipo] = {"qtd": bucket["qtd"], "sla": sla_avg}

    docs = {"meses": meses_out, "ordem": ordem, "tipos": tipos}

    # ---- rev_arte ----
    ranking = []
    cliente_counter = Counter()
    cliente_prod = defaultdict(set)
    marca_counter = Counter()
    marca_prod = defaultdict(set)
    total_revisoes = 0
    distribuicao = Counter()

    for (prod, cliente), entradas in arte_group.items():
        n = len(entradas)
        versoes = [e["versao"] for e in entradas if e["versao"] is not None]
        max_versao = max(versoes) if versoes else n
        # "n_revisoes" considera o numero total de correcoes (entradas) de
        # Correcao de Arte registradas para este produto/cliente.
        n_revisoes = n
        total_revisoes += n_revisoes
        datas = [e["entrada"] for e in entradas if e["entrada"]]
        primeira = min(datas).isoformat() if datas else None
        ultima = max(datas).isoformat() if datas else None
        marca = next((e["marca"] for e in entradas if e["marca"]), "")

        ranking.append({
            "produto": prod,
            "cliente": cliente,
            "marca": marca,
            "n_revisoes": n_revisoes,
            "max_versao": max_versao,
            "primeira_entrada": primeira,
            "ultima_entrada": ultima,
        })

        if cliente:
            cliente_counter[cliente] += n_revisoes
            cliente_prod[cliente].add(prod)
        if marca:
            marca_counter[marca] += n_revisoes
            marca_prod[marca].add(prod)

        bucket = "5+" if n_revisoes >= 5 else str(n_revisoes)
        distribuicao[bucket] += 1

    ranking.sort(key=lambda r: r["n_revisoes"], reverse=True)

    ranking_cliente = [
        {"nome": nome, "n_revisoes": n, "n_produtos": len(cliente_prod[nome])}
        for nome, n in cliente_counter.most_common()
    ]
    ranking_marca = [
        {"nome": nome, "n_revisoes": n, "n_produtos": len(marca_prod[nome])}
        for nome, n in marca_counter.most_common()
    ]

    total_produtos_cliente = len(arte_group)
    media_revisoes = round(total_revisoes / total_produtos_cliente, 2) if total_produtos_cliente else 0.0
    com_retrabalho = sum(1 for r in ranking if r["n_revisoes"] > 1)
    pct_com_retrabalho = round(100 * com_retrabalho / total_produtos_cliente, 2) if total_produtos_cliente else 0.0

    rev_arte = {
        "total_produtos_cliente": total_produtos_cliente,
        "total_revisoes": total_revisoes,
        "media_revisoes": media_revisoes,
        "pct_com_retrabalho": pct_com_retrabalho,
        "distribuicao": dict(distribuicao),
        "ranking": ranking,
        "ranking_cliente": ranking_cliente,
        "ranking_marca": ranking_marca,
    }

    return docs, rev_arte


# --------------------------------------------------------------------------
# 5) Notificacoes vigentes (Gestao Notificações)
# --------------------------------------------------------------------------

def _top_counter(counter, top=10, other_label="Outros"):
    items = counter.most_common()
    if len(items) <= top:
        return [[k, v] for k, v in items]
    head = items[: top - 1]
    tail_sum = sum(v for _, v in items[top - 1:])
    out = [[k, v] for k, v in head]
    out.append([other_label, tail_sum])
    return out


def _extract_notif_vig(wb):
    sheet_name = "Gestão Notificações" if "Gestão Notificações" in wb.sheetnames else "Notificações"
    ws = wb[sheet_name]
    header_idx, header = _find_header_row(
        ws, ["Produto", "Sabor", "Embalagem", "Peso Líquido", "Validade", "Marcas", "Cliente",
             "Data vigência", "processo", "Status"]
    )
    c_produto = _col_index(header, "Produto")
    c_sabor = _col_index(header, "Sabor")
    c_embalagem = _col_index(header, "Embalagem")
    c_peso = _col_index(header, "Peso Líquido")
    c_validade = _col_index(header, "Validade")
    c_marca = _col_index(header, "Marcas")
    c_cliente = _col_index(header, "Cliente")
    c_vigencia = _col_index(header, "Data vigência")
    c_processo = _col_index(header, "processo")
    c_status = _col_index(header, "Status")

    lista = []
    embalagem_c = Counter()
    marca_c = Counter()
    cliente_c = Counter()
    peso_c = Counter()
    por_produto = defaultdict(lambda: {"sabores": set(), "pesos": set(), "n": 0})
    all_sabores = set()
    all_pesos = set()

    seen_keys = set()
    for row in _iter_data_rows(ws, header_idx):
        status = row[c_status] if c_status < len(row) else None

        # Considera tambem produtos novos ainda sem status "ANUIDO" (ex.:
        # "Aguardando", "No aguardo" ou status em branco), para que apareçam
        # no indicador "por produto" assim que entram na planilha.
        produto = row[c_produto] if c_produto < len(row) else None
        if not produto:
            continue
        sabor = row[c_sabor] if c_sabor < len(row) else None
        embalagem = row[c_embalagem] if c_embalagem < len(row) else None
        peso_raw = row[c_peso] if c_peso < len(row) else None
        peso = f"{int(peso_raw)} g" if isinstance(peso_raw, (int, float)) else (str(peso_raw) if peso_raw else None)
        validade = row[c_validade] if c_validade < len(row) else None
        marca = normalize_name(row[c_marca]) if c_marca < len(row) else None
        cliente = normalize_name(row[c_cliente]) if c_cliente < len(row) else None
        vigencia = _to_date(row[c_vigencia]) if c_vigencia < len(row) else None
        processo = row[c_processo] if c_processo < len(row) else None

        produto_s = str(produto).strip()

        # A identidade de uma notificacao vigente e dada pela combinacao de
        # numero de processo + produto + sabor + embalagem + peso liquido.
        # Linhas duplicadas dessa combinacao (ex.: multiplas atualizacoes de
        # status para o mesmo registro) sao contadas uma unica vez.
        dedup_key = (
            str(processo).strip() if processo else "",
            produto_s,
            str(sabor).strip() if sabor else "",
            str(embalagem).strip() if embalagem else "",
            peso or "",
        )
        if dedup_key in seen_keys:
            continue
        seen_keys.add(dedup_key)

        lista.append({
            "produto": produto_s,
            "sabor": str(sabor).strip() if sabor else "",
            "embalagem": str(embalagem).strip() if embalagem else "",
            "peso": peso or "",
            "marca": marca or "",
            "cliente": cliente or "",
            "validade_meses": int(validade) if isinstance(validade, (int, float)) else None,
            "data_vigencia": vigencia.isoformat() if vigencia else None,
            "processo": str(processo).strip() if processo else "",
        })

        if embalagem:
            embalagem_c[str(embalagem).strip()] += 1
        if marca:
            marca_c[marca] += 1
        if cliente:
            cliente_c[cliente] += 1
        if peso:
            peso_c[peso] += 1

        pp = por_produto[produto_s]
        if sabor:
            pp["sabores"].add(str(sabor).strip())
            all_sabores.add(str(sabor).strip())
        if peso:
            pp["pesos"].add(peso)
            all_pesos.add(peso)
        pp["n"] += 1

    por_produto_out = []
    for produto, info in por_produto.items():
        por_produto_out.append({
            "produto": produto,
            "n_sabores": len(info["sabores"]),
            "sabores": sorted(info["sabores"]),
            "n_pesos": len(info["pesos"]),
            "pesos": sorted(info["pesos"]),
            "n_notificacoes": info["n"],
        })
    por_produto_out.sort(key=lambda x: x["n_notificacoes"], reverse=True)

    return {
        "total": len(lista),
        "produtos": len(por_produto),
        "embalagem": _top_counter(embalagem_c, top=5),
        "marca": _top_counter(marca_c, top=11),
        "cliente": _top_counter(cliente_c, top=11),
        "peso": _top_counter(peso_c, top=12),
        "lista": lista,
        "por_produto": por_produto_out,
        "total_sabores": len(all_sabores),
        "total_pesos": len(all_pesos),
    }


# --------------------------------------------------------------------------
# 6) Gastos Clientes (usa aba 'Gastos Reais')
# --------------------------------------------------------------------------

def _extract_gastos_clientes(wb):
    ws = wb["Gastos Reais"]
    header_idx, header = _find_header_row(
        ws, ["Status", "Valor", "Descrição", "Cliente", "Nome Abreviado", "Pagamento"]
    )
    c_status = _col_index(header, "Status")
    c_valor = _col_index(header, "Valor")
    c_descricao = _col_index(header, "Descrição")
    c_cliente = _col_index(header, "Cliente")
    c_lab = _col_index(header, "Nome Abreviado")
    c_pagamento = _col_index(header, "Pagamento")

    registros = []
    ranking_c = Counter()
    detalhe = defaultdict(lambda: {"pago": 0.0, "a_vencer": 0.0, "n": 0})
    total_pago = 0.0
    total_a_vencer = 0.0

    for row in _iter_data_rows(ws, header_idx):
        cliente = row[c_cliente] if c_cliente < len(row) else None
        if not cliente:
            continue
        status = row[c_status] if c_status < len(row) else None
        valor = _to_number(row[c_valor]) if c_valor < len(row) else None
        descricao = row[c_descricao] if c_descricao < len(row) else None
        lab = row[c_lab] if c_lab < len(row) else None
        pagamento = _to_date(row[c_pagamento]) if c_pagamento < len(row) else None

        cliente_n = normalize_name(cliente)
        status_s = str(status).strip() if status else ""

        registros.append({
            "cliente": cliente_n,
            "valor": round(valor, 2) if valor is not None else None,
            "status": status_s,
            "lab": str(lab).strip() if lab else "",
            "data_pagamento": pagamento.isoformat() if pagamento else None,
            "descricao": str(descricao).strip() if descricao else "",
        })

        if valor is not None:
            ranking_c[cliente_n] += valor
            if status_s.lower() == "pago":
                detalhe[cliente_n]["pago"] += valor
                total_pago += valor
            elif status_s.lower() == "a vencer":
                detalhe[cliente_n]["a_vencer"] += valor
                total_a_vencer += valor
            detalhe[cliente_n]["n"] += 1

    ranking = [[k, round(v, 2)] for k, v in ranking_c.most_common()]
    ranking_detalhe = {
        k: {"pago": round(v["pago"], 2), "a_vencer": round(v["a_vencer"], 2), "n": v["n"]}
        for k, v in detalhe.items()
    }

    return {
        "fonte": "Gastos Reais",
        "registros": registros,
        "ranking": ranking,
        "ranking_detalhe": ranking_detalhe,
        "total_pago": round(total_pago, 2),
        "total_a_vencer": round(total_a_vencer, 2),
        "n_registros": len(registros),
    }


# --------------------------------------------------------------------------
# build()
# --------------------------------------------------------------------------

def build(xlsx_path, today=None):
    """Extrai todos os KPIs da planilha `xlsx_path` e retorna um dict
    JSON-serializavel no formato consumido pelo dashboard HTML.

    `today` (date, opcional) e usado para calcular os "proximos laudos"
    aguardados em `estab`. Se omitido, usa a data atual do sistema.
    """
    import openpyxl

    if today is None:
        today = date.today()
    elif isinstance(today, datetime):
        today = today.date()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    cotacoes = _extract_cotacoes(wb)
    fluxo = _extract_fluxo(wb)
    estab, qualidade_lab, lab_sla = _extract_estab_qualidade_prob(wb, today)
    docs, rev_arte = _extract_docs_and_revarte(wb)
    notif_vig = _extract_notif_vig(wb)
    gastos_clientes = _extract_gastos_clientes(wb)

    agora = datetime.now()
    meta = {
        "atualizado_em": agora.strftime("%d/%m/%Y %H:%M"),
        "atualizado_em_iso": agora.isoformat(),
    }

    return {
        "cotacoes": cotacoes,
        "fluxo": fluxo,
        "estab": estab,
        "qualidade_lab": qualidade_lab,
        "lab_sla": lab_sla,
        "docs": docs,
        "notif_vig": notif_vig,
        "gastos_clientes": gastos_clientes,
        "rev_arte": rev_arte,
        "meta": meta,
    }


if __name__ == "__main__":
    import sys
    import json

    path = sys.argv[1] if len(sys.argv) > 1 else "LISTA MESTRE DE PRODUTOS - PRONUTRITION.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "kpis_all_new.json"
    d = build(path)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)
    print(f"Gerado {out}")
